#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
HR Data Validation Script
Validates data inconsistencies between incentive CSV file and team_structure_updated.json
"""

import pandas as pd
import json
import os
import sys
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import warnings
import glob
warnings.filterwarnings('ignore')

class HRDataValidator:
    def __init__(self, month, year):
        self.month = month
        self.year = year
        self.errors = []
        self.warnings = []
        self.base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

        # Set file paths
        self.csv_path = os.path.join(self.base_path, 'input_files', f'{year}년 {month}월 인센티브 지급 세부 정보.csv')
        self.json_path = os.path.join(self.base_path, 'HR info', 'team_structure_updated.json')
        self.output_dir = os.path.join(self.base_path, 'error_review')

        # Create output directory
        os.makedirs(self.output_dir, exist_ok=True)

        # Load data
        self.csv_data = None
        self.json_data = None

    def load_data(self):
        """Load CSV and JSON data."""
        try:
            # Load CSV data
            print(f"📂 Loading CSV file: {self.csv_path}")
            self.csv_data = pd.read_csv(self.csv_path, encoding='utf-8-sig')
            print(f"   ✅ {len(self.csv_data)} employee data loaded successfully")

            # Load JSON data
            print(f"📂 Loading JSON file: {self.json_path}")
            with open(self.json_path, 'r', encoding='utf-8') as f:
                self.json_data = json.load(f)
            print(f"   ✅ {len(self.json_data['positions'])} position definitions loaded successfully")

            return True

        except FileNotFoundError as e:
            print(f"❌ File not found: {e}")
            return False
        except Exception as e:
            print(f"❌ Error occurred while loading data: {e}")
            return False

    def validate_position_mapping(self):
        """Validate position mapping."""
        print("\n🔍 Validating position mapping...")

        # Create position mapping from JSON
        json_positions = {}
        for pos in self.json_data['positions']:
            key = (
                pos.get('position_1st', '').strip().upper(),
                pos.get('position_2nd', '').strip().upper(),
                pos.get('position_3rd', '').strip().upper()
            )
            json_positions[key] = {
                'final_code': pos.get('final_code', ''),
                'team_name': pos.get('team_name', ''),
                'role_category': pos.get('role_category', ''),
                'role_type': pos.get('role_type', '')
            }

        # Validate CSV data
        mismatches = []
        for idx, row in self.csv_data.iterrows():
            csv_key = (
                str(row.get('QIP POSITION 1ST  NAME', '')).strip().upper(),
                str(row.get('QIP POSITION 2ND  NAME', '')).strip().upper(),
                str(row.get('QIP POSITION 3RD  NAME', '')).strip().upper()
            )

            employee_info = {
                'Employee No': row.get('Employee No', ''),
                'Full Name': row.get('Full Name', ''),
                'Position 1st': csv_key[0],
                'Position 2nd': csv_key[1],
                'Position 3rd': csv_key[2],
                'CSV_Final_Code': row.get('FINAL QIP POSITION NAME CODE', ''),
                'CSV_Role_Type': row.get('ROLE TYPE STD', ''),
                'Error_Type': '',
                'Expected_Values': '',
                'Actual_Values': '',
                'Reason of Review Required': ''
            }

            if csv_key not in json_positions:
                # Case: Position not found in JSON
                employee_info['Error_Type'] = 'Position Not Found in JSON'
                employee_info['Expected_Values'] = 'Position definition in JSON'
                employee_info['Actual_Values'] = f"{csv_key[0]} / {csv_key[1]} / {csv_key[2]}"
                employee_info['Reason of Review Required'] = 'Position combination not defined in team_structure_updated.json. May be new position or typo'
                mismatches.append(employee_info)
            else:
                # Case: Position exists but values differ
                json_info = json_positions[csv_key]
                errors = []

                # Validate Final Code
                if str(row.get('FINAL QIP POSITION NAME CODE', '')).strip() != json_info['final_code']:
                    errors.append('Final Code Mismatch')
                    employee_info['Expected_Values'] = f"Final Code: {json_info['final_code']}"
                    employee_info['Actual_Values'] = f"Final Code: {row.get('FINAL QIP POSITION NAME CODE', '')}"
                    employee_info['Reason of Review Required'] = 'Final Code does not match JSON definition. Code update required'

                # Validate Role Type
                if str(row.get('ROLE TYPE STD', '')).strip() != json_info['role_type']:
                    errors.append('Role Type Mismatch')
                    if employee_info['Expected_Values']:
                        employee_info['Expected_Values'] += f", Role Type: {json_info['role_type']}"
                        employee_info['Actual_Values'] += f", Role Type: {row.get('ROLE TYPE STD', '')}"
                        if employee_info['Reason of Review Required']:
                            employee_info['Reason of Review Required'] += ' / Role Type (TYPE-1/2/3) classification mismatch. Affects incentive calculation'
                        else:
                            employee_info['Reason of Review Required'] = 'Role Type (TYPE-1/2/3) classification mismatch. Affects incentive calculation'
                    else:
                        employee_info['Expected_Values'] = f"Role Type: {json_info['role_type']}"
                        employee_info['Actual_Values'] = f"Role Type: {row.get('ROLE TYPE STD', '')}"
                        employee_info['Reason of Review Required'] = 'Role Type (TYPE-1/2/3) classification mismatch. Affects incentive calculation'

                if errors:
                    employee_info['Error_Type'] = ', '.join(errors)
                    mismatches.append(employee_info)

        print(f"   ✅ Validation completed: {len(mismatches)} mismatches found")
        return mismatches

    def validate_role_type_consistency(self):
        """Validate Role Type consistency for same position."""
        print("\n🔍 Validating Role Type consistency...")

        position_types = {}
        inconsistencies = []

        for idx, row in self.csv_data.iterrows():
            position_1st = str(row.get('QIP POSITION 1ST  NAME', '')).strip().upper()
            role_type = str(row.get('ROLE TYPE STD', '')).strip()

            if position_1st not in position_types:
                position_types[position_1st] = set()

            position_types[position_1st].add(role_type)

        # Find cases where same position_1st has multiple role_types
        for position, types in position_types.items():
            if len(types) > 1:
                # Find all employees with this position
                affected_employees = self.csv_data[
                    self.csv_data['QIP POSITION 1ST  NAME'].str.strip().str.upper() == position
                ]

                for idx, row in affected_employees.iterrows():
                    inconsistencies.append({
                        'Employee No': row.get('Employee No', ''),
                        'Full Name': row.get('Full Name', ''),
                        'Position 1st': position,
                        'Current Role Type': row.get('ROLE TYPE STD', ''),
                        'All Role Types for Position': ', '.join(sorted(types)),
                        'Issue': 'Multiple Role Types for Same Position',
                        'Reason of Review Required': f'Multiple TYPEs ({", ".join(sorted(types))}) mixed for same position ({position}). Need to unify'
                    })

        print(f"   ✅ Validation completed: {len(inconsistencies)} consistency issues found")
        return inconsistencies

    def validate_duplicate_codes(self):
        """Validate duplicate Final Codes."""
        print("\n🔍 Validating duplicate Final Codes...")

        code_positions = {}
        duplicates = []

        # Collect positions by Final Code from JSON
        for pos in self.json_data['positions']:
            code = pos.get('final_code', '')
            if code:
                if code not in code_positions:
                    code_positions[code] = []
                code_positions[code].append({
                    'position_1st': pos.get('position_1st', ''),
                    'position_2nd': pos.get('position_2nd', ''),
                    'position_3rd': pos.get('position_3rd', ''),
                    'team': pos.get('team_name', ''),
                    'role_type': pos.get('role_type', '')
                })

        # Find duplicate codes
        for code, positions in code_positions.items():
            if len(positions) > 1:
                for pos in positions:
                    duplicates.append({
                        'Final Code': code,
                        'Position 1st': pos['position_1st'],
                        'Position 2nd': pos['position_2nd'],
                        'Position 3rd': pos['position_3rd'],
                        'Team': pos['team'],
                        'Role Type': pos['role_type'],
                        'Issue': f'Code used by {len(positions)} positions',
                        'Reason of Review Required': f'Final Code ({code}) used by {len(positions)} positions (duplicate). Need to reassign unique codes'
                    })

        print(f"   ✅ Validation completed: {len(duplicates)} duplicate codes found")
        return duplicates

    def save_to_excel(self, mismatches, inconsistencies, duplicates):
        """Save validation results to Excel file (keep only latest)."""
        # Delete previous validation files
        old_pattern = os.path.join(self.output_dir, f'hr_data_validation_{self.year}_{self.month}_*.xlsx')
        old_files = glob.glob(old_pattern)
        for old_file in old_files:
            try:
                os.remove(old_file)
                print(f"🗑️  Previous validation file deleted: {os.path.basename(old_file)}")
            except Exception as e:
                print(f"⚠️  Failed to delete validation file: {old_file} - {e}")

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = os.path.join(self.output_dir, f'hr_data_validation_{self.year}_{self.month}_{timestamp}.xlsx')

        print(f"💾 Creating Excel file: {os.path.basename(output_file)}")

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # 1. Position mapping mismatches
            if mismatches:
                df_mismatches = pd.DataFrame(mismatches)
                df_mismatches.to_excel(writer, sheet_name='Position Mismatches', index=False)

                # Apply style
                worksheet = writer.sheets['Position Mismatches']
                self._apply_excel_style(worksheet, len(mismatches))

            # 2. Role Type consistency issues
            if inconsistencies:
                df_inconsistencies = pd.DataFrame(inconsistencies)
                df_inconsistencies.to_excel(writer, sheet_name='Role Type Inconsistencies', index=False)

                worksheet = writer.sheets['Role Type Inconsistencies']
                self._apply_excel_style(worksheet, len(inconsistencies))

            # 3. Duplicate Final Codes
            if duplicates:
                df_duplicates = pd.DataFrame(duplicates)
                df_duplicates.to_excel(writer, sheet_name='Duplicate Final Codes', index=False)

                worksheet = writer.sheets['Duplicate Final Codes']
                self._apply_excel_style(worksheet, len(duplicates))

            # 4. Summary sheet
            summary_data = {
                'Validation Type': [
                    'Position Mismatches',
                    'Role Type Inconsistencies',
                    'Duplicate Final Codes',
                    'Total Issues'
                ],
                'Count': [
                    len(mismatches),
                    len(inconsistencies),
                    len(duplicates),
                    len(mismatches) + len(inconsistencies) + len(duplicates)
                ],
                'Description': [
                    'Position information mismatch between CSV and JSON',
                    'Multiple role types exist for same position',
                    'Same final code used by multiple positions',
                    'Total number of issues found'
                ],
                'Impact on Dashboard': [
                    'Team assignment error, displayed as Team Unidentified',
                    'Potential incentive amount calculation errors',
                    'Data mapping confusion, report accuracy degradation',
                    'Dashboard reliability degradation'
                ],
                'Recommended Action': [
                    'Add missing positions to team_structure_updated.json',
                    'Unify Role Type (choose one of TYPE-1/2/3) for same position',
                    'Reassign unique Final Code to each position',
                    'Review data consistency and re-run'
                ]
            }
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary', index=False)

            worksheet = writer.sheets['Summary']
            self._apply_excel_style(worksheet, len(summary_data['Validation Type']))

        print(f"   ✅ Excel file saved successfully")
        return output_file

    def _apply_excel_style(self, worksheet, row_count):
        """Apply style to Excel worksheet."""
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Apply header row style
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Apply data row style
        for row in worksheet.iter_rows(min_row=2, max_row=row_count+1):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center')

        # Auto-adjust column width
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def run_validation(self):
        """Run complete validation process."""
        print(f"\n{'='*60}")
        print(f"HR Data Validation Started - {self.year} {self.month}")
        print(f"{'='*60}")

        # Load data
        if not self.load_data():
            return False

        # Perform validation
        mismatches = self.validate_position_mapping()
        inconsistencies = self.validate_role_type_consistency()
        duplicates = self.validate_duplicate_codes()

        # Save results
        if mismatches or inconsistencies or duplicates:
            output_file = self.save_to_excel(mismatches, inconsistencies, duplicates)

            print(f"\n📊 Validation Results Summary:")
            print(f"   • Position mismatches: {len(mismatches)} issues")
            print(f"   • Role Type consistency issues: {len(inconsistencies)} issues")
            print(f"   • Duplicate Final Codes: {len(duplicates)} issues")
            print(f"   • Total issues: {len(mismatches) + len(inconsistencies) + len(duplicates)}")
            print(f"\n📄 See detailed results in:")
            print(f"   {output_file}")

            return True
        else:
            print(f"\n✅ All data is valid. No inconsistencies found.")
            return True

def main():
    """Main execution function"""
    # Process command line arguments
    if len(sys.argv) > 2:
        month = int(sys.argv[1])
        year = int(sys.argv[2])
    else:
        # Default: current month
        from datetime import datetime
        now = datetime.now()
        month = now.month
        year = now.year

    # Run validation
    validator = HRDataValidator(month, year)
    success = validator.run_validation()

    if not success:
        sys.exit(1)

if __name__ == "__main__":
    main()
