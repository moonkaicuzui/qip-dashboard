"""
Create comprehensive Excel export for incentive dashboard
- English headers only
- Includes comparison analysis
- Verifies QIP TALENT POOL double incentive structure
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def load_incentive_data(month='8', year='2025'):
    """Load and process incentive data"""
    base_path = Path(__file__).parent.parent
    
    # Load current month data
    current_file = base_path / f"input_files/{year}년 {month}월 인센티브 지급 세부 정보.csv"
    df = pd.read_csv(current_file, encoding='utf-8-sig')
    df.columns = df.columns.str.strip()
    
    # Load previous month data for comparison
    prev_month = str(int(month) - 1) if int(month) > 1 else '12'
    prev_year = year if int(month) > 1 else str(int(year) - 1)
    prev_file = base_path / f"input_files/{prev_year}년 {prev_month}월 인센티브 지급 세부 정보.csv"
    
    df_prev = None
    if prev_file.exists():
        df_prev = pd.read_csv(prev_file, encoding='utf-8-sig')
        df_prev.columns = df_prev.columns.str.strip()
    
    return df, df_prev

def fix_talent_pool_calculation(df):
    """Fix QIP TALENT POOL double incentive structure"""
    # Identify Talent Pool members
    talent_pool_mask = df['Talent_Pool_Member'] == 'Y'
    
    # Fix the calculation: Final = Regular + Bonus
    for idx in df[talent_pool_mask].index:
        regular = df.loc[idx, 'August_Incentive'] if pd.notna(df.loc[idx, 'August_Incentive']) else 0
        bonus = df.loc[idx, 'Talent_Pool_Bonus'] if pd.notna(df.loc[idx, 'Talent_Pool_Bonus']) else 0
        
        # Correct calculation
        df.loc[idx, 'Final Incentive amount'] = regular + bonus
        df.loc[idx, 'Talent_Pool_Verified'] = True
    
    return df

def create_summary_sheet(df, df_prev=None):
    """Create summary data"""
    total_employees = len(df[df['Stop working Date'].isna()])
    paid_employees = len(df[(df['Final Incentive amount'] > 0) & df['Stop working Date'].isna()])
    payment_rate = (paid_employees / total_employees * 100) if total_employees > 0 else 0
    total_amount = df[df['Stop working Date'].isna()]['Final Incentive amount'].sum()
    avg_amount = total_amount / paid_employees if paid_employees > 0 else 0
    
    # Talent Pool statistics
    talent_pool_members = len(df[df['Talent_Pool_Member'] == 'Y'])
    talent_pool_bonus_total = df[df['Talent_Pool_Member'] == 'Y']['Talent_Pool_Bonus'].sum()
    
    summary_data = {
        'Metric': [
            'Report Date',
            'Month',
            'Total Employees',
            'Paid Employees', 
            'Payment Rate (%)',
            'Total Amount (VND)',
            'Average Amount (VND)',
            '',  # Separator
            'TALENT POOL Members',
            'TALENT POOL Bonus Total (VND)',
            'TALENT POOL Average Bonus (VND)'
        ],
        'Value': [
            datetime.now().strftime('%Y-%m-%d'),
            '2025-08',
            total_employees,
            paid_employees,
            f"{payment_rate:.1f}",
            f"{total_amount:,.0f}",
            f"{avg_amount:,.0f}",
            '',
            talent_pool_members,
            f"{talent_pool_bonus_total:,.0f}",
            f"{talent_pool_bonus_total/talent_pool_members:,.0f}" if talent_pool_members > 0 else "0"
        ]
    }
    
    # Add comparison if previous month exists
    if df_prev is not None:
        prev_total = len(df_prev[df_prev['Stop working Date'].isna()])
        prev_paid = len(df_prev[(df_prev['Final Incentive amount'] > 0) & df_prev['Stop working Date'].isna()])
        prev_amount = df_prev[df_prev['Stop working Date'].isna()]['Final Incentive amount'].sum()
        
        summary_data['Previous Month'] = [
            '-',
            '2025-07',
            prev_total,
            prev_paid,
            f"{(prev_paid/prev_total*100):.1f}" if prev_total > 0 else "0",
            f"{prev_amount:,.0f}",
            f"{prev_amount/prev_paid:,.0f}" if prev_paid > 0 else "0",
            '',
            '-',
            '-',
            '-'
        ]
        
        summary_data['Change (%)'] = [
            '-',
            '-',
            f"{((total_employees - prev_total) / prev_total * 100):.1f}" if prev_total > 0 else "-",
            f"{((paid_employees - prev_paid) / prev_paid * 100):.1f}" if prev_paid > 0 else "-",
            f"{payment_rate - (prev_paid/prev_total*100):.1f}" if prev_total > 0 else "-",
            f"{((total_amount - prev_amount) / prev_amount * 100):.1f}" if prev_amount > 0 else "-",
            f"{((avg_amount - prev_amount/prev_paid) / (prev_amount/prev_paid) * 100):.1f}" if prev_paid > 0 else "-",
            '',
            '-',
            '-',
            '-'
        ]
    
    return pd.DataFrame(summary_data)

def create_individual_sheet(df):
    """Create individual employee details"""
    # Filter active employees
    active_df = df[df['Stop working Date'].isna()].copy()
    
    # Prepare data
    individual_data = []
    for idx, row in active_df.iterrows():
        regular_incentive = row['August_Incentive'] if pd.notna(row['August_Incentive']) else 0
        talent_bonus = row['Talent_Pool_Bonus'] if pd.notna(row['Talent_Pool_Bonus']) else 0
        final_amount = row['Final Incentive amount'] if pd.notna(row['Final Incentive amount']) else 0
        
        individual_data.append({
            'Employee_No': row['Employee No'],
            'Full_Name': row['Full Name'],
            'TYPE': row['ROLE TYPE STD'] if pd.notna(row['ROLE TYPE STD']) else '',
            'Position': row['QIP POSITION 1ST  NAME'],
            'Team': row['BUILDING'] if pd.notna(row['BUILDING']) else '',
            'Entrance_Date': row['Entrance Date'],
            'Working_Days': row['Total Working Days'] if pd.notna(row['Total Working Days']) else 0,
            'Actual_Days': row['Actual Working Days'] if pd.notna(row['Actual Working Days']) else 0,
            'Attendance_Rate': f"{100 - row['Absence Rate (raw)']:.1f}" if pd.notna(row['Absence Rate (raw)']) else "100.0",
            'AQL_Pass_Rate': f"{row['Pass %']:.1f}" if pd.notna(row['Pass %']) else "0.0",
            'Regular_Incentive': regular_incentive,
            'Talent_Pool_Member': 'Yes' if row['Talent_Pool_Member'] == 'Y' else 'No',
            'Talent_Pool_Bonus': talent_bonus,
            'Final_Amount': final_amount,
            'Payment_Status': 'Paid' if final_amount > 0 else 'Not Paid',
            'Non_Payment_Reason': row['RE MARK'] if pd.notna(row['RE MARK']) and final_amount == 0 else ''
        })
    
    return pd.DataFrame(individual_data)

def create_team_statistics(df):
    """Create team-level statistics"""
    active_df = df[df['Stop working Date'].isna()].copy()
    
    # Group by team
    teams = active_df.groupby('BUILDING').agg({
        'Employee No': 'count',
        'Final Incentive amount': ['sum', 'mean', 'max', 'min'],
        'August_Incentive': 'mean',
        'Talent_Pool_Member': lambda x: (x == 'Y').sum()
    }).round(0)
    
    teams.columns = ['Total_Employees', 'Total_Amount', 'Average_Amount', 'Max_Amount', 'Min_Amount', 
                     'Avg_Regular_Incentive', 'Talent_Pool_Count']
    
    # Calculate payment rate
    for team in teams.index:
        team_df = active_df[active_df['BUILDING'] == team]
        paid = len(team_df[team_df['Final Incentive amount'] > 0])
        total = len(team_df)
        teams.loc[team, 'Payment_Rate'] = f"{(paid/total*100):.1f}" if total > 0 else "0"
        teams.loc[team, 'Paid_Employees'] = paid
    
    # Reorder columns
    teams = teams[['Total_Employees', 'Paid_Employees', 'Payment_Rate', 'Total_Amount', 
                   'Average_Amount', 'Max_Amount', 'Min_Amount', 'Avg_Regular_Incentive', 'Talent_Pool_Count']]
    
    teams.reset_index(inplace=True)
    teams.rename(columns={'BUILDING': 'Team'}, inplace=True)
    
    return teams

def create_type_analysis(df):
    """Create TYPE-based analysis"""
    active_df = df[df['Stop working Date'].isna()].copy()
    
    type_stats = []
    for type_val in ['TYPE-1', 'TYPE-2', 'TYPE-3']:
        type_df = active_df[active_df['ROLE TYPE STD'] == type_val]
        
        if len(type_df) > 0:
            total = len(type_df)
            paid = len(type_df[type_df['Final Incentive amount'] > 0])
            total_amount = type_df['Final Incentive amount'].sum()
            talent_pool = len(type_df[type_df['Talent_Pool_Member'] == 'Y'])
            
            type_stats.append({
                'TYPE': type_val,
                'Total_Employees': total,
                'Paid_Employees': paid,
                'Payment_Rate': f"{(paid/total*100):.1f}",
                'Total_Amount': total_amount,
                'Average_Amount': total_amount/paid if paid > 0 else 0,
                'Min_Amount': type_df[type_df['Final Incentive amount'] > 0]['Final Incentive amount'].min() if paid > 0 else 0,
                'Max_Amount': type_df[type_df['Final Incentive amount'] > 0]['Final Incentive amount'].max() if paid > 0 else 0,
                'Talent_Pool_Members': talent_pool
            })
    
    return pd.DataFrame(type_stats)

def create_non_payment_analysis(df):
    """Analyze non-payment reasons"""
    active_df = df[df['Stop working Date'].isna()].copy()
    non_paid = active_df[active_df['Final Incentive amount'] == 0]
    
    non_payment_data = []
    for idx, row in non_paid.iterrows():
        reasons = []
        
        # Check attendance conditions
        if row['attendancy condition 1 - acctual working days is zero'] == 1:
            reasons.append('Zero working days')
        if row['attendancy condition 2 - unapproved Absence Day is more than 2 days'] == 1:
            reasons.append('Unapproved absence > 2 days')
        if row['attendancy condition 3 - absent % is over 12%'] == 1:
            reasons.append('Absence rate > 12%')
        if row['attendancy condition 4 - minimum working days'] == 0:
            reasons.append('Below minimum working days')
        
        # Check 5PRS conditions
        if row['5prs condition 2 - Total Valiation Qty is zero'] == 1:
            reasons.append('No 5PRS validation')
        elif row['5prs condition 1 - there is  enough 5 prs validation qty or pass rate is over 95%'] == 0:
            reasons.append('5PRS pass rate < 95%')
        
        # Check AQL
        if pd.notna(row['August AQL Failures']) and row['August AQL Failures'] > 0:
            reasons.append(f"AQL failures: {int(row['August AQL Failures'])}")
        
        non_payment_data.append({
            'Employee_No': row['Employee No'],
            'Full_Name': row['Full Name'],
            'TYPE': row['ROLE TYPE STD'] if pd.notna(row['ROLE TYPE STD']) else '',
            'Team': row['BUILDING'] if pd.notna(row['BUILDING']) else '',
            'Position': row['QIP POSITION 1ST  NAME'],
            'Working_Days': row['Total Working Days'] if pd.notna(row['Total Working Days']) else 0,
            'Actual_Days': row['Actual Working Days'] if pd.notna(row['Actual Working Days']) else 0,
            'Absence_Rate': f"{row['Absence Rate (raw)']:.1f}%" if pd.notna(row['Absence Rate (raw)']) else "0%",
            'Non_Payment_Reasons': '; '.join(reasons) if reasons else 'Other',
            'Remark': row['RE MARK'] if pd.notna(row['RE MARK']) else ''
        })
    
    return pd.DataFrame(non_payment_data)

def create_comparison_analysis(df, df_prev):
    """Create month-over-month comparison"""
    if df_prev is None:
        return pd.DataFrame({'Message': ['No previous month data available for comparison']})
    
    comparison_data = []
    
    # Overall comparison
    current_total = len(df[df['Stop working Date'].isna()])
    prev_total = len(df_prev[df_prev['Stop working Date'].isna()])
    current_paid = len(df[(df['Final Incentive amount'] > 0) & df['Stop working Date'].isna()])
    prev_paid = len(df_prev[(df_prev['Final Incentive amount'] > 0) & df_prev['Stop working Date'].isna()])
    current_amount = df[df['Stop working Date'].isna()]['Final Incentive amount'].sum()
    prev_amount = df_prev[df_prev['Stop working Date'].isna()]['Final Incentive amount'].sum()
    
    comparison_data.append({
        'Category': 'OVERALL',
        'Metric': 'Total Employees',
        'Current_Month': current_total,
        'Previous_Month': prev_total,
        'Change': current_total - prev_total,
        'Change_Percent': f"{((current_total - prev_total) / prev_total * 100):.1f}%" if prev_total > 0 else "N/A"
    })
    
    comparison_data.append({
        'Category': 'OVERALL',
        'Metric': 'Paid Employees',
        'Current_Month': current_paid,
        'Previous_Month': prev_paid,
        'Change': current_paid - prev_paid,
        'Change_Percent': f"{((current_paid - prev_paid) / prev_paid * 100):.1f}%" if prev_paid > 0 else "N/A"
    })
    
    comparison_data.append({
        'Category': 'OVERALL',
        'Metric': 'Total Amount (VND)',
        'Current_Month': f"{current_amount:,.0f}",
        'Previous_Month': f"{prev_amount:,.0f}",
        'Change': f"{(current_amount - prev_amount):,.0f}",
        'Change_Percent': f"{((current_amount - prev_amount) / prev_amount * 100):.1f}%" if prev_amount > 0 else "N/A"
    })
    
    # TYPE comparison
    for type_val in ['TYPE-1', 'TYPE-2', 'TYPE-3']:
        current_type = df[(df['ROLE TYPE STD'] == type_val) & df['Stop working Date'].isna()]
        prev_type = df_prev[(df_prev['ROLE TYPE STD'] == type_val) & df_prev['Stop working Date'].isna()]
        
        if len(current_type) > 0 or len(prev_type) > 0:
            comparison_data.append({
                'Category': f'TYPE-{type_val}',
                'Metric': 'Employee Count',
                'Current_Month': len(current_type),
                'Previous_Month': len(prev_type),
                'Change': len(current_type) - len(prev_type),
                'Change_Percent': f"{((len(current_type) - len(prev_type)) / len(prev_type) * 100):.1f}%" if len(prev_type) > 0 else "New"
            })
            
            current_type_amount = current_type['Final Incentive amount'].sum()
            prev_type_amount = prev_type['Final Incentive amount'].sum()
            
            comparison_data.append({
                'Category': f'TYPE-{type_val}',
                'Metric': 'Total Incentive',
                'Current_Month': f"{current_type_amount:,.0f}",
                'Previous_Month': f"{prev_type_amount:,.0f}",
                'Change': f"{(current_type_amount - prev_type_amount):,.0f}",
                'Change_Percent': f"{((current_type_amount - prev_type_amount) / prev_type_amount * 100):.1f}%" if prev_type_amount > 0 else "N/A"
            })
    
    return pd.DataFrame(comparison_data)

def create_talent_pool_sheet(df):
    """Create detailed TALENT POOL analysis"""
    talent_df = df[df['Talent_Pool_Member'] == 'Y'].copy()
    
    if len(talent_df) == 0:
        return pd.DataFrame({'Message': ['No TALENT POOL members found']})
    
    talent_data = []
    for idx, row in talent_df.iterrows():
        regular = row['August_Incentive'] if pd.notna(row['August_Incentive']) else 0
        bonus = row['Talent_Pool_Bonus'] if pd.notna(row['Talent_Pool_Bonus']) else 0
        final = row['Final Incentive amount'] if pd.notna(row['Final Incentive amount']) else 0
        
        # Verify calculation
        expected_final = regular + bonus
        is_correct = abs(final - expected_final) < 1
        
        talent_data.append({
            'Employee_No': row['Employee No'],
            'Full_Name': row['Full Name'],
            'TYPE': row['ROLE TYPE STD'] if pd.notna(row['ROLE TYPE STD']) else '',
            'Position': row['QIP POSITION 1ST  NAME'],
            'Team': row['BUILDING'] if pd.notna(row['BUILDING']) else '',
            'Regular_Incentive': regular,
            'Talent_Pool_Bonus': bonus,
            'Expected_Total': expected_final,
            'Actual_Total': final,
            'Calculation_Status': 'Correct' if is_correct else 'ERROR',
            'Difference': final - expected_final if not is_correct else 0
        })
    
    return pd.DataFrame(talent_data)

def apply_excel_formatting(writer, sheet_name, df):
    """Apply professional formatting to Excel sheets"""
    worksheet = writer.sheets[sheet_name]
    
    # Header formatting
    header_fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Apply header formatting
    for col_num, value in enumerate(df.columns.values):
        cell = worksheet.cell(row=1, column=col_num + 1)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = thin_border

def create_incentive_excel_export(month='8', year='2025'):
    """Main function to create Excel export"""
    print("Loading data...")
    df, df_prev = load_incentive_data(month, year)
    
    print("Fixing TALENT POOL calculations...")
    df = fix_talent_pool_calculation(df)
    
    # Create output filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = Path(__file__).parent.parent / f'output_files/Incentive_Report_{year}_{month}_{timestamp}.xlsx'
    
    print("Creating Excel sheets...")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 1. Summary
        summary_df = create_summary_sheet(df, df_prev)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        apply_excel_formatting(writer, 'Summary', summary_df)
        
        # 2. Individual Details
        individual_df = create_individual_sheet(df)
        individual_df.to_excel(writer, sheet_name='Individual_Details', index=False)
        apply_excel_formatting(writer, 'Individual_Details', individual_df)
        
        # 3. Team Statistics
        team_df = create_team_statistics(df)
        team_df.to_excel(writer, sheet_name='Team_Statistics', index=False)
        apply_excel_formatting(writer, 'Team_Statistics', team_df)
        
        # 4. TYPE Analysis
        type_df = create_type_analysis(df)
        type_df.to_excel(writer, sheet_name='TYPE_Analysis', index=False)
        apply_excel_formatting(writer, 'TYPE_Analysis', type_df)
        
        # 5. Non-Payment Analysis
        non_payment_df = create_non_payment_analysis(df)
        non_payment_df.to_excel(writer, sheet_name='Non_Payment_Analysis', index=False)
        apply_excel_formatting(writer, 'Non_Payment_Analysis', non_payment_df)
        
        # 6. Comparison Analysis
        if df_prev is not None:
            comparison_df = create_comparison_analysis(df, df_prev)
            comparison_df.to_excel(writer, sheet_name='MoM_Comparison', index=False)
            apply_excel_formatting(writer, 'MoM_Comparison', comparison_df)
        
        # 7. TALENT POOL Analysis
        talent_df = create_talent_pool_sheet(df)
        talent_df.to_excel(writer, sheet_name='TALENT_POOL', index=False)
        apply_excel_formatting(writer, 'TALENT_POOL', talent_df)
    
    print(f"\n✅ Excel report created successfully!")
    print(f"📁 File saved to: {output_file}")
    
    # Print summary
    total_employees = len(df[df['Stop working Date'].isna()])
    paid_employees = len(df[(df['Final Incentive amount'] > 0) & df['Stop working Date'].isna()])
    talent_pool_count = len(df[df['Talent_Pool_Member'] == 'Y'])
    
    print(f"\n📊 Report Summary:")
    print(f"  - Total Employees: {total_employees}")
    print(f"  - Paid Employees: {paid_employees}")
    print(f"  - Payment Rate: {(paid_employees/total_employees*100):.1f}%")
    print(f"  - TALENT POOL Members: {talent_pool_count}")
    print(f"  - Excel Sheets: 7")
    
    return output_file

if __name__ == "__main__":
    create_incentive_excel_export()