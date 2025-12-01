#!/usr/bin/env python3
"""
간편 Validation 리포트 생성기
CSV 파일의 기존 조건 평가 데이터를 활용하여 validation 리포트 생성
"""

import pandas as pd
import sys
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_validation_report(month: str, year: int):
    """CSV 데이터를 기반으로 validation 리포트 생성"""

    print("=" * 80)
    print(f"📊 Validation 리포트 생성 - {year}년 {month}")
    print("=" * 80)
    print()

    # V9.1 → V9.0 → V8.02 순서로 확인 - 통일된 fallback 패턴 (2025-12-01)
    csv_path_v91 = Path(f"output_files/output_QIP_incentive_{month}_{year}_Complete_V9.1_Complete.csv")
    csv_path_v9 = Path(f"output_files/output_QIP_incentive_{month}_{year}_Complete_V9.0_Complete.csv")
    csv_path_v8 = Path(f"output_files/output_QIP_incentive_{month}_{year}_Complete_V8.02_Complete.csv")

    if csv_path_v91.exists():
        csv_path = csv_path_v91
    elif csv_path_v9.exists():
        csv_path = csv_path_v9
    elif csv_path_v8.exists():
        csv_path = csv_path_v8
    else:
        csv_path = csv_path_v91  # For error message

    if not csv_path.exists():
        print(f"❌ CSV 파일을 찾을 수 없습니다: {csv_path}")
        return 1

    print(f"📂 CSV 로드: {csv_path}")
    df = pd.read_csv(csv_path)
    print(f"   ✅ {len(df)}명의 직원 데이터 로드")
    print()

    # Month number mapping
    month_map = {
        'january': 1, 'february': 2, 'march': 3, 'april': 4,
        'may': 5, 'june': 6, 'july': 7, 'august': 8,
        'september': 9, 'october': 10, 'november': 11, 'december': 12
    }
    month_num = month_map.get(month.lower(), 9)

    # Active 직원만 필터링 (Dashboard와 동일한 로직)
    # 해당 월 이전에 퇴사한 사람만 제외 (해당 월 중 퇴사는 포함)
    if 'Stop working Date' in df.columns:
        month_start = pd.Timestamp(year=year, month=month_num, day=1)

        def is_active(row):
            stop_date = row['Stop working Date']
            if pd.isna(stop_date) or stop_date == '':
                return True
            try:
                # Dashboard와 동일한 파싱 (dayfirst 파라미터 없음 - pandas 기본값 사용)
                resignation_date = pd.to_datetime(stop_date, errors='coerce')
                if pd.notna(resignation_date):
                    # 해당 월 시작일 이전에 퇴사한 경우만 제외
                    return resignation_date >= month_start
                return True  # 날짜 변환 실패 시 포함
            except:
                return True  # 예외 발생 시 포함

        active_df = df[df.apply(is_active, axis=1)].copy()
        resigned_before_month = len(df) - len(active_df)
        print(f"📊 Active 직원: {len(active_df)}명 ({month} {year} 이전 퇴사자 {resigned_before_month}명 제외)")
    else:
        active_df = df.copy()
        print(f"📊 전체 직원: {len(active_df)}명")
    print()

    # Validation 리포트 폴더 생성
    report_dir = Path("validation_reports")
    report_dir.mkdir(exist_ok=True)

    # Excel 파일 생성
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = report_dir / f"VALIDATION_REPORT_{month}_{year}_{timestamp}.xlsx"

    wb = openpyxl.Workbook()

    # 1. Summary Sheet (전체 df도 함께 전달)
    create_summary_sheet(wb, active_df, month, year, df)

    # 2. Condition Details Sheet
    create_condition_details_sheet(wb, active_df)

    # 3. Incentive Distribution Sheet
    create_incentive_distribution_sheet(wb, active_df)

    # 4. Failed Conditions Sheet
    create_failed_conditions_sheet(wb, active_df)

    # 기본 Sheet 제거
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # 저장
    wb.save(report_path)

    print("=" * 80)
    print(f"✅ Validation 리포트 생성 완료!")
    print(f"📄 파일: {report_path}")
    print("=" * 80)

    return 0

def create_summary_sheet(wb, df, month, year, full_df=None):
    """Summary Sheet 생성"""
    ws = wb.active
    ws.title = "Summary"

    # 헤더
    ws['A1'] = f"QIP Incentive Validation Report - {year}년 {month}"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells('A1:D1')

    # 생성 정보
    ws['A2'] = f"생성 시각: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(size=10, italic=True)

    row = 4

    # 전체 통계
    ws[f'A{row}'] = "📊 전체 통계 (인센티브 평가 대상: TYPE-1, TYPE-2)"
    ws[f'A{row}'].font = Font(size=14, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.merge_cells(f'A{row}:D{row}')

    row += 1
    stats = [
        ("전체 직원 수", len(df)),
        ("인센티브 수령자", len(df[df['Final Incentive amount'] > 0])),
        ("인센티브 미수령자", len(df[df['Final Incentive amount'] == 0])),
        ("총 지급액 (VND)", f"{df['Final Incentive amount'].sum():,.0f}"),
        ("평균 인센티브 (VND)", f"{df[df['Final Incentive amount'] > 0]['Final Incentive amount'].mean():,.0f}"),
    ]

    for label, value in stats:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

    row += 1

    # 전체 직원 무단결근 현황 (TYPE-3 포함)
    if full_df is not None:
        ws[f'A{row}'] = "⚠️ 전체 직원 무단결근 현황 (TYPE-3 포함)"
        ws[f'A{row}'].font = Font(size=14, bold=True, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        ws.merge_cells(f'A{row}:D{row}')

        row += 1

        # Active 직원만 필터 (같은 로직 적용)
        month_map = {
            'january': 1, 'february': 2, 'march': 3, 'april': 4,
            'may': 5, 'june': 6, 'july': 7, 'august': 8,
            'september': 9, 'october': 10, 'november': 11, 'december': 12
        }
        month_num = month_map.get(month.lower(), 9)
        month_start = pd.Timestamp(year=year, month=month_num, day=1)

        def is_active(row):
            stop_date = row['Stop working Date']
            if pd.isna(stop_date) or stop_date == '':
                return True
            try:
                resignation_date = pd.to_datetime(stop_date, errors='coerce')
                if pd.notna(resignation_date):
                    return resignation_date >= month_start
                return True
            except:
                return True

        active_full_df = full_df[full_df.apply(is_active, axis=1)].copy()

        # 무단결근 > 2일 직원 통계
        if 'Unapproved Absences' in active_full_df.columns:
            # NaN 제외하고 계산
            unapproved_with_data = active_full_df[active_full_df['Unapproved Absences'].notna()].copy()
            excessive_absences = unapproved_with_data[unapproved_with_data['Unapproved Absences'] > 2]

            absence_stats = [
                ("전체 Active 직원", len(active_full_df)),
                ("출결 데이터 있는 직원", len(unapproved_with_data)),
                ("무단결근 > 2일 직원", len(excessive_absences)),
                ("무단결근 <= 2일 직원", len(unapproved_with_data) - len(excessive_absences)),
            ]

            for label, value in absence_stats:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value

                # 무단결근 > 2일 행을 빨간색으로 강조
                if "무단결근 > 2일" in label:
                    ws[f'A{row}'].font = Font(bold=True, color="C0392B")
                    ws[f'B{row}'].font = Font(bold=True, color="C0392B")
                    ws[f'B{row}'].fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
                else:
                    ws[f'A{row}'].font = Font(bold=True)

                row += 1

            # TYPE별 무단결근 통계
            row += 1
            ws[f'A{row}'] = "TYPE별 무단결근 > 2일 분포"
            ws[f'A{row}'].font = Font(bold=True, italic=True)
            row += 1

            for role_type in sorted(active_full_df['ROLE TYPE STD'].unique()):
                type_df = active_full_df[active_full_df['ROLE TYPE STD'] == role_type]
                type_with_data = type_df[type_df['Unapproved Absences'].notna()]
                type_excessive = type_with_data[type_with_data['Unapproved Absences'] > 2]

                ws[f'A{row}'] = f"  {role_type}"
                ws[f'B{row}'] = f"{len(type_excessive)}명"
                row += 1

        row += 1

    # TYPE별 통계
    ws[f'A{row}'] = "📋 TYPE별 인센티브 통계"
    ws[f'A{row}'].font = Font(size=14, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    ws.merge_cells(f'A{row}:D{row}')

    row += 1
    ws[f'A{row}'] = "TYPE"
    ws[f'B{row}'] = "직원 수"
    ws[f'C{row}'] = "수령자"
    ws[f'D{row}'] = "총 지급액"

    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row}'].font = Font(bold=True)
        ws[f'{col}{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    row += 1

    for role_type in sorted(df['ROLE TYPE STD'].unique()):
        type_df = df[df['ROLE TYPE STD'] == role_type]
        receiving = type_df[type_df['Final Incentive amount'] > 0]

        ws[f'A{row}'] = role_type
        ws[f'B{row}'] = len(type_df)
        ws[f'C{row}'] = len(receiving)
        ws[f'D{row}'] = f"{type_df['Final Incentive amount'].sum():,.0f}"
        row += 1

    # 컬럼 너비 조정
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25

def create_condition_details_sheet(wb, df):
    """Condition Details Sheet 생성"""
    ws = wb.create_sheet("Condition Details")

    # 헤더
    ws['A1'] = "조건별 통과/실패 현황"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells('A1:F1')

    # 컬럼 헤더
    headers = ["조건", "통과", "실패", "N/A", "통과율 (%)", "비고"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # 조건별 통계 (조건명, 컬럼명, 비고)
    conditions = [
        ("C1: 출근율 >= 88%", "cond_1_attendance_rate", "TYPE-3 제외"),
        ("C2: 무단결근 <= 2일", "cond_2_unapproved_absence", "TYPE-3 제외"),
        ("C3: 실제근무일 > 0", "cond_3_actual_working_days", "TYPE-3 제외"),
        ("C4: 최소근무일 >= 12일", "cond_4_minimum_days", "TYPE-3 제외"),
        ("C5: AQL 개인 실패 = 0", "cond_5_aql_personal_failure", "전체 대상 (AQL 검사자)"),
        ("C6: AQL 3개월 연속 실패 없음", "cond_6_aql_continuous", "전체 대상 (AQL 검사자)"),
        ("C7: AQL 팀/구역 3개월 연속 실패 없음", "cond_7_aql_team_area", "전체 대상 (AQL 검사자)"),
        ("C8: 구역 Reject율 < 3%", "cond_8_area_reject", "전체 대상 (AQL 검사자)"),
        ("C9: 5PRS 통과율 >= 95%", "cond_9_5prs_pass_rate", "TYPE-3 제외 (5PRS 검사자)"),
        ("C10: 5PRS 검사량 >= 100", "cond_10_5prs_inspection_qty", "TYPE-3 제외 (5PRS 검사자)"),
    ]

    row = 4
    for cond_name, cond_col, note in conditions:
        if cond_col not in df.columns:
            continue

        passed = (df[cond_col] == 'PASS').sum()
        failed = (df[cond_col] == 'FAIL').sum()
        na = (df[cond_col] == 'N/A').sum()
        total_applicable = passed + failed
        pass_rate = (passed / total_applicable * 100) if total_applicable > 0 else 0

        ws[f'A{row}'] = cond_name
        ws[f'B{row}'] = passed
        ws[f'C{row}'] = failed
        ws[f'D{row}'] = na
        ws[f'E{row}'] = f"{pass_rate:.1f}%"
        ws[f'F{row}'] = note

        # 통과율에 따라 색상 지정
        if pass_rate >= 90:
            ws[f'E{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif pass_rate >= 70:
            ws[f'E{row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            ws[f'E{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        row += 1

    # 컬럼 너비 조정
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30

def create_incentive_distribution_sheet(wb, df):
    """Incentive Distribution Sheet 생성"""
    ws = wb.create_sheet("Incentive Distribution")

    # 헤더
    ws['A1'] = "인센티브 분포 현황"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells('A1:D1')

    # 금액대별 분포
    ws['A3'] = "금액대"
    ws['B3'] = "직원 수"
    ws['C3'] = "비율 (%)"
    ws['D3'] = "총 지급액 (VND)"

    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}3'].font = Font(bold=True)
        ws[f'{col}3'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # 금액대별 집계
    bins = [0, 1, 50000, 100000, 200000, 300000, 500000, 1000000, float('inf')]
    labels = ['0 VND', '1-50K', '50K-100K', '100K-200K', '200K-300K', '300K-500K', '500K-1M', '1M+']

    df['amount_range'] = pd.cut(df['Final Incentive amount'], bins=bins, labels=labels, right=False)

    row = 4
    total_count = len(df)

    for label in labels:
        range_df = df[df['amount_range'] == label]
        count = len(range_df)
        pct = (count / total_count * 100) if total_count > 0 else 0
        total_amt = range_df['Final Incentive amount'].sum()

        ws[f'A{row}'] = label
        ws[f'B{row}'] = count
        ws[f'C{row}'] = f"{pct:.1f}%"
        ws[f'D{row}'] = f"{total_amt:,.0f}"
        row += 1

    # 컬럼 너비 조정
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25

def create_failed_conditions_sheet(wb, df):
    """Failed Conditions Sheet 생성"""
    ws = wb.create_sheet("Failed Conditions")

    # 헤더
    ws['A1'] = "조건 미충족 직원 목록 (인센티브 0 VND)"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws.merge_cells('A1:F1')

    # 조건 미충족 직원 필터링
    failed_df = df[(df['Final Incentive amount'] == 0) & (df['conditions_pass_rate'] < 100)].copy()

    if len(failed_df) == 0:
        ws['A3'] = "✅ 모든 직원이 조건을 충족하거나 해당없음입니다!"
        ws['A3'].font = Font(size=12, bold=True, color="70AD47")
        return

    # 컬럼 헤더
    headers = ["Employee No", "Full Name", "Position", "TYPE", "통과율 (%)", "실패한 조건"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # 데이터 입력
    row = 4
    for _, emp in failed_df.iterrows():
        # 실패한 조건 찾기
        failed_conds = []
        for i in range(1, 11):
            cond_col = f'cond_{i}_attendance_rate' if i == 1 else \
                      f'cond_{i}_unapproved_absence' if i == 2 else \
                      f'cond_{i}_actual_working_days' if i == 3 else \
                      f'cond_{i}_minimum_days' if i == 4 else \
                      f'cond_{i}_aql_personal_failure' if i == 5 else \
                      f'cond_{i}_aql_continuous' if i == 6 else \
                      f'cond_{i}_aql_team_area' if i == 7 else \
                      f'cond_{i}_area_reject' if i == 8 else \
                      f'cond_{i}_5prs_pass_rate' if i == 9 else \
                      f'cond_{i}_5prs_inspection_qty'

            if cond_col in emp.index and emp[cond_col] == 'FAIL':
                failed_conds.append(f'C{i}')

        ws[f'A{row}'] = emp['Employee No']
        ws[f'B{row}'] = emp['Full Name']
        ws[f'C{row}'] = emp['QIP POSITION 1ST  NAME']
        ws[f'D{row}'] = emp['ROLE TYPE STD']
        ws[f'E{row}'] = f"{emp.get('conditions_pass_rate', 0):.1f}%"
        ws[f'F{row}'] = ', '.join(failed_conds)

        # 통과율이 낮을수록 빨간색
        pass_rate = emp.get('conditions_pass_rate', 0)
        if pass_rate < 50:
            ws[f'E{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif pass_rate < 80:
            ws[f'E{row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        row += 1

    # 컬럼 너비 조정
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20

def main():
    if len(sys.argv) < 3:
        print("Usage: python generate_simple_validation_report.py <month> <year>")
        print("Example: python generate_simple_validation_report.py september 2025")
        sys.exit(1)

    month = sys.argv[1].lower()
    year = int(sys.argv[2])

    result = create_validation_report(month, year)
    sys.exit(result)

if __name__ == "__main__":
    main()
